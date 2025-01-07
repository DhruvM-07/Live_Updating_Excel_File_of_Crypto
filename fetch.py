import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
import time

# API Endpoint and Parameters
API_URL = "https://api.coingecko.com/api/v3/coins/markets"
PARAMS = {
    "vs_currency": "usd",
    "order": "market_cap_desc",
    "per_page": 50,
    "page": 1
}

# Function to fetch cryptocurrency data
def fetch_crypto_data():
    response = requests.get(API_URL, params=PARAMS)
    if response.status_code == 200:
        data = response.json()
        return [
            {
                "Name": coin["name"],
                "Symbol": coin["symbol"].upper(),
                "Price (USD)": coin["current_price"],
                "Market Cap (USD)": coin["market_cap"],
                "24h Volume (USD)": coin["total_volume"],
                "Price Change 24h (%)": coin["price_change_percentage_24h"],
            }
            for coin in data
        ]
    else:
        print(f"API Error: {response.status_code}")
        return []

# Function to analyze cryptocurrency data
def analyze_data(data):
    # Convert to DataFrame
    df = pd.DataFrame(data)

    # Identify Top 5 by Market Cap
    top_5 = df.nlargest(5, "Market Cap (USD)")

    # Average Price of Top 50
    average_price = df["Price (USD)"].mean()

    # Highest and Lowest 24h % Change
    max_change = df.loc[df["Price Change 24h (%)"].idxmax()]
    min_change = df.loc[df["Price Change 24h (%)"].idxmin()]

    return top_5, average_price, max_change, min_change

# Function to update Excel sheet
def update_excel(data, analysis):
    file_name = "Crypto_Live_Data.xlsx"
    try:
        wb = load_workbook(file_name)
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Live Crypto Data"
        sheet.append(["Name", "Symbol", "Price (USD)", "Market Cap (USD)", 
                      "24h Volume (USD)", "Price Change 24h (%)"])

    # Clear old data
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)

    # Write new data
    for row in data:
        sheet.append(list(row.values()))

    # Write analysis to a separate sheet
    analysis_sheet_name = "Analysis"
    if analysis_sheet_name not in wb.sheetnames:
        wb.create_sheet(title=analysis_sheet_name)
    analysis_sheet = wb[analysis_sheet_name]

    # Clear old analysis data
    if analysis_sheet.max_row > 0:
        analysis_sheet.delete_rows(1, analysis_sheet.max_row)

    # Write analysis results
    analysis_sheet.append(["Metric", "Value"])
    analysis_sheet.append(["Average Price", f"${analysis[1]:.2f}"])
    analysis_sheet.append([
        "Highest 24h Change",
        f"{analysis[2]['Name']} ({analysis[2]['Price Change 24h (%)']:.2f}%)"
    ])
    analysis_sheet.append([
        "Lowest 24h Change",
        f"{analysis[3]['Name']} ({analysis[3]['Price Change 24h (%)']:.2f}%)"
    ])

    # Save the workbook
    wb.save(file_name)
    print(f"Excel file '{file_name}' updated successfully!")

# Continuous update loop
def main():
    while True:
        try:
            # Fetch data
            data = fetch_crypto_data()
            if data:
                # Analyze data
                analysis = analyze_data(data)
                # Update Excel
                update_excel(data, analysis)
                print("Excel updated successfully!")
            else:
                print("No data fetched from API.")
        except Exception as e:
            print(f"Error during update: {e}")

        # Wait for 5 minutes before the next update
        time.sleep(300)

# Entry point
if __name__ == "__main__":
    main()
